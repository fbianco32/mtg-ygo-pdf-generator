import os
from pathlib import Path
from strategies.YGOStrategy import YGOHandler
from strategies.MTGStrategy import MTGHandler
from strategies.DIGIMONStrategy import DIGIMONHandler
from DeckTypes import YGO, MTG, DIGIMON
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

input_fodler = Path.cwd()/'input'
output_folder = Path.cwd()/'output'

typeDirectoryInputPaths = {
    YGO: input_fodler/'YGO',
    MTG: input_fodler/'MTG',
    DIGIMON: input_fodler/'DIGIMON'
}
##Sheet config
header_row_number = 1
received_list_column_number = 4
received_list_column_text = "received"
missing_list_column_number = 5
missing_list_column_text = "missing"
missing_total_column_number = 6
missing_total_column_text = "total_missing"
type_column_number = 7
type_column_text = "type"
##

def createYGOHandler(type):
    return YGOHandler(typeDirectoryInputPaths.get(type))
def createMTGHandler(type):
    return MTGHandler(typeDirectoryInputPaths.get(type))
def createDIGIMONHandler(type):
    return DIGIMONHandler(typeDirectoryInputPaths.get(type))
def default():
    return None

handlerSwitch = {
    YGO: createYGOHandler(YGO),
    MTG: createMTGHandler(MTG),
    DIGIMON: createDIGIMONHandler(DIGIMON)
}

def makeDirsIfNotExists():
    Path.mkdir(input_fodler,exist_ok=True)
    Path.mkdir(output_folder,exist_ok=True)
    Path.mkdir(typeDirectoryInputPaths[YGO],exist_ok=True)
    Path.mkdir(typeDirectoryInputPaths[MTG],exist_ok=True)
    Path.mkdir(typeDirectoryInputPaths[DIGIMON],exist_ok=True)

def getHandler(type):
    return handlerSwitch.get(type, default)

def addRecievedColumn(sheet):
    sheet.cell(row=header_row_number, column=received_list_column_number).value = received_list_column_text #should be on a global config variable

    for row in sheet.iter_rows(min_row=header_row_number+1, max_row=len(sheet["A"]), min_col=received_list_column_number, max_col=received_list_column_number, values_only=False):
        current_cell = row[0]
        max_value = current_cell.offset(row=0, column=-1).value
        valid_options = '"'+", ".join(map(str, range(max_value+1)))+'"' #excel inner workings wonkiness
        rule = DataValidation(type='list', formula1=valid_options, allow_blank=False)
        sheet.add_data_validation(rule)
        current_cell.value = 0
        rule.add(current_cell)

def addMissingColumns(sheet):
    sheet.cell(row=header_row_number, column=missing_list_column_number).value = missing_list_column_text 
    sheet.cell(row=header_row_number, column=missing_total_column_number).value = missing_total_column_text

    for row in sheet.iter_rows(min_row=header_row_number+1, max_row=len(sheet["A"]), min_col=missing_list_column_number, max_col=missing_list_column_number, values_only=False):
        current_cell = row[0]
        expected_value_cell = current_cell.offset(row=0, column=-2)
        recieved_value_cell = current_cell.offset(row=0, column=-1)
        current_cell.value = f"={expected_value_cell.coordinate}-{recieved_value_cell.coordinate}"
        error_fill = PatternFill(start_color='F84F31',end_color='F84F31',fill_type='solid')
        ok_fill = PatternFill(start_color='23C552',end_color='23C552',fill_type='solid')
        sheet.conditional_formatting.add(f'{current_cell.coordinate}', CellIsRule(operator='greaterThan', formula=[0], fill=error_fill))
        sheet.conditional_formatting.add(f'{current_cell.coordinate}', CellIsRule(operator='equal', formula=[0], fill=ok_fill))
    
    add_missing_from = sheet.cell(row=2, column= missing_list_column_number).coordinate
    add_missing_to = sheet.cell(row=len(sheet["E"]), column= missing_list_column_number).coordinate #hardcoded, should change 
    sheet.cell(row=header_row_number+1, column=missing_total_column_number).value = f"=SUM({add_missing_from}:{add_missing_to})"


def adjustStyle(sheet):
    for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']: #hardcoded, should change 
        max_width = 0
        for row_number in range(1, sheet.max_row + 1):
            if len(str(sheet[f'{letter}{row_number}'].value)) > max_width:
                max_width = len(str(sheet[f'{letter}{row_number}'].value))
        sheet.column_dimensions[letter].width = max_width + 3

    sheet.freeze_panes = 'A2' #hardcoded, should change 

def addTypeColumn(sheet, type):
    sheet.cell(row=1, column=type_column_number).value = type_column_text
    sheet.cell(row=2, column=type_column_number).value = type
               

def transfromGenericListsToChecklists(genericLists):
    workbook = Workbook()
    del workbook['Sheet']
    for genericList in genericLists:
        sheet = workbook.create_sheet(genericList.getName())
        rows = dataframe_to_rows(genericList.getDataframe(), index=False, header=True)
        for row in rows:
            sheet.append(row)
        addRecievedColumn(sheet)
        addMissingColumns(sheet)
        addTypeColumn(sheet, genericList.getType())
        adjustStyle(sheet)
    return workbook

def getChecklistName():
    while True:
        try:
            value = str(input("Please enter the decks owner name: "))
        except ValueError:
            print("Sorry, your response must be a non-empty string")
            continue
        break
    return value

def main():
    handlers = []
    print("-----Initializing-----")
    print("Creating input directories")
    makeDirsIfNotExists()
    print("You may now place decks into created input directories")
    print("-----Initializing completed-----")
    checklistName = getChecklistName()
    print("-----Processing started-----")
    for type in typeDirectoryInputPaths:
        if os.listdir(typeDirectoryInputPaths.get(type)) != []:
            handlers.append(getHandler(type))

    genericLists = []
    for handler in handlers:
        print(f"Processing {handler.getType()} decks")
        print("Getting decks")
        decklists = handler.getDecklists()
        print("Transforming decks to generic lists")
        handler.transformDecklistsToGenericLists(decklists)
        genericLists.extend(handler.getGenericLists())
    
    print("Building Excel sheets")
    checklist = transfromGenericListsToChecklists(genericLists)
    print("Saving file")
    checklist.save(filename=output_folder/f"{checklistName}.xlsx")

    print("-----Processing finished-----")

if __name__ == "__main__":
    main()